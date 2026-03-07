"""
Gelişmiş Arama Penceresi - Comprehensive Search Dialog
Proje_Q - JCL Veri Yönetim Sistemi
v0.6.0
"""
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel,
                             QLineEdit, QComboBox, QCheckBox, QPushButton,
                             QTableWidget, QTableWidgetItem, QGroupBox,
                             QTabWidget, QWidget, QSpinBox, QDateEdit,
                             QListWidget, QMessageBox, QHeaderView,
                             QTextEdit, QSplitter)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor
import json
import os
from datetime import datetime


class AdvancedSearchDialog(QDialog):
    """Gelişmiş Arama Penceresi"""
    
    def __init__(self, parent, db_manager):
        super().__init__(parent)
        self.parent = parent
        self.db_manager = db_manager
        self.search_history_file = 'config/search_history.json'
        self.favorites_file = 'config/search_favorites.json'
        
        # Logger'ı import et
        from utils.logger import app_logger
        self.logger = app_logger
        
        # Arama geçmişi ve favoriler
        self.search_history = self.load_search_history()
        self.favorites = self.load_favorites()
        
        self.logger.info("Gelişmiş arama penceresi açıldı")
        
        self.init_ui()
        self.load_combo_data()
        
    def init_ui(self):
        """UI'yi oluştur"""
        self.setWindowTitle("🔍 Gelişmiş Arama - Proje_Q")
        self.setGeometry(150, 100, 1400, 800)
        
        # Ana layout
        layout = QVBoxLayout(self)
        
        # Splitter - Sol: Arama Kriterleri, Sağ: Sonuçlar
        splitter = QSplitter(Qt.Horizontal)
        
        # Sol Panel - Arama Kriterleri
        left_panel = self.create_search_criteria_panel()
        splitter.addWidget(left_panel)
        
        # Sağ Panel - Sonuçlar
        right_panel = self.create_results_panel()
        splitter.addWidget(right_panel)
        
        # Sol %40, Sağ %60
        splitter.setStretchFactor(0, 40)
        splitter.setStretchFactor(1, 60)
        
        layout.addWidget(splitter)
        
        # Alt butonlar
        bottom_buttons = self.create_bottom_buttons()
        layout.addLayout(bottom_buttons)
        
    def create_search_criteria_panel(self):
        """Arama kriterleri paneli"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Tab widget
        self.criteria_tabs = QTabWidget()
        
        # Tab 1: Temel Arama
        basic_tab = self.create_basic_search_tab()
        self.criteria_tabs.addTab(basic_tab, "📝 Temel Arama")
        
        # Tab 2: Sayısal Filtreler
        numeric_tab = self.create_numeric_filters_tab()
        self.criteria_tabs.addTab(numeric_tab, "📊 Sayısal Filtreler")
        
        # Tab 3: Gelişmiş Metin
        advanced_tab = self.create_advanced_text_tab()
        self.criteria_tabs.addTab(advanced_tab, "🔤 Gelişmiş Metin")
        
        # Tab 4: Toplu JCL Sorgulama
        bulk_tab = self.create_bulk_jcl_tab()
        self.criteria_tabs.addTab(bulk_tab, "📋 Toplu JCL Sorgulama")
        
        # Tab 5: Geçmiş & Favoriler
        history_tab = self.create_history_favorites_tab()
        self.criteria_tabs.addTab(history_tab, "⭐ Geçmiş & Favoriler")
        
        layout.addWidget(self.criteria_tabs)
        
        # Arama butonları
        search_buttons = QHBoxLayout()
        
        self.btn_search = QPushButton("🔍 Ara")
        self.btn_search.clicked.connect(self.perform_search)
        self.btn_search.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;")
        
        self.btn_reset = QPushButton("🔄 Sıfırla")
        self.btn_reset.clicked.connect(self.reset_criteria)
        
        self.btn_save_search = QPushButton("💾 Aramayı Kaydet")
        self.btn_save_search.clicked.connect(self.save_current_search)
        
        search_buttons.addWidget(self.btn_search)
        search_buttons.addWidget(self.btn_reset)
        search_buttons.addWidget(self.btn_save_search)
        
        layout.addLayout(search_buttons)
        
        return widget
    
    def create_basic_search_tab(self):
        """Temel arama tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # JCL Arama
        jcl_group = QGroupBox("JCL Arama")
        jcl_layout = QVBoxLayout(jcl_group)
        
        self.jcl_input = QLineEdit()
        self.jcl_input.setPlaceholderText("JCL adı girin (virgül ile ayırarak birden fazla, wildcard: PONT*)")
        
        self.cb_jcl_wildcard = QCheckBox("🔍 Wildcard kullan (* ve ?)")
        self.cb_jcl_wildcard.setChecked(True)
        
        self.cb_jcl_case_sensitive = QCheckBox("Aa Büyük/küçük harf duyarlı")
        
        jcl_layout.addWidget(QLabel("JCL Adı:"))
        jcl_layout.addWidget(self.jcl_input)
        jcl_layout.addWidget(self.cb_jcl_wildcard)
        jcl_layout.addWidget(self.cb_jcl_case_sensitive)
        
        layout.addWidget(jcl_group)
        
        # Ekip ve Ay
        filters_group = QGroupBox("Filtreler")
        filters_layout = QVBoxLayout(filters_group)
        
        # Ekip
        ekip_row = QHBoxLayout()
        ekip_row.addWidget(QLabel("Sorumlu Ekip:"))
        self.ekip_combo = QComboBox()
        ekip_row.addWidget(self.ekip_combo)
        filters_layout.addLayout(ekip_row)
        
        # Ay
        ay_row = QHBoxLayout()
        ay_row.addWidget(QLabel("Ay:"))
        self.ay_combo = QComboBox()
        ay_row.addWidget(self.ay_combo)
        filters_layout.addLayout(ay_row)
        
        layout.addWidget(filters_group)
        
        # Rapor Tipi
        rapor_group = QGroupBox("Rapor Tipi")
        rapor_layout = QVBoxLayout(rapor_group)
        
        self.cb_hatali = QCheckBox("✅ Hatalı İşler")
        self.cb_hatali.setChecked(True)
        
        self.cb_uzun = QCheckBox("⏱️ Uzun Süren İşler")
        self.cb_uzun.setChecked(True)
        
        rapor_layout.addWidget(self.cb_hatali)
        rapor_layout.addWidget(self.cb_uzun)
        
        layout.addWidget(rapor_group)
        
        layout.addStretch()
        
        return widget
    
    def create_numeric_filters_tab(self):
        """Sayısal filtreler tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Hatalı Sayı Aralığı
        hatali_group = QGroupBox("📊 Hatalı İş Sayısı Aralığı")
        hatali_layout = QHBoxLayout(hatali_group)
        
        self.cb_hatali_filter = QCheckBox("Etkinleştir")
        hatali_layout.addWidget(self.cb_hatali_filter)
        
        hatali_layout.addWidget(QLabel("Min:"))
        self.hatali_min = QSpinBox()
        self.hatali_min.setMaximum(999999)
        self.hatali_min.setEnabled(False)
        hatali_layout.addWidget(self.hatali_min)
        
        hatali_layout.addWidget(QLabel("Max:"))
        self.hatali_max = QSpinBox()
        self.hatali_max.setMaximum(999999)
        self.hatali_max.setValue(1000)
        self.hatali_max.setEnabled(False)
        hatali_layout.addWidget(self.hatali_max)
        
        self.cb_hatali_filter.stateChanged.connect(
            lambda: self.toggle_numeric_filter(self.cb_hatali_filter, [self.hatali_min, self.hatali_max])
        )
        
        layout.addWidget(hatali_group)
        
        # Süre Aralığı
        sure_group = QGroupBox("⏱️ Süre Aralığı (Dakika)")
        sure_layout = QHBoxLayout(sure_group)
        
        self.cb_sure_filter = QCheckBox("Etkinleştir")
        sure_layout.addWidget(self.cb_sure_filter)
        
        sure_layout.addWidget(QLabel("Min:"))
        self.sure_min = QSpinBox()
        self.sure_min.setMaximum(999999)
        self.sure_min.setEnabled(False)
        sure_layout.addWidget(self.sure_min)
        
        sure_layout.addWidget(QLabel("Max:"))
        self.sure_max = QSpinBox()
        self.sure_max.setMaximum(999999)
        self.sure_max.setValue(240)
        self.sure_max.setEnabled(False)
        sure_layout.addWidget(self.sure_max)
        
        self.cb_sure_filter.stateChanged.connect(
            lambda: self.toggle_numeric_filter(self.cb_sure_filter, [self.sure_min, self.sure_max])
        )
        
        layout.addWidget(sure_group)
        
        # Tarih Aralığı
        tarih_group = QGroupBox("📅 Tarih Aralığı")
        tarih_layout = QVBoxLayout(tarih_group)
        
        self.cb_tarih_filter = QCheckBox("Etkinleştir")
        tarih_layout.addWidget(self.cb_tarih_filter)
        
        tarih_row = QHBoxLayout()
        tarih_row.addWidget(QLabel("Başlangıç:"))
        self.tarih_baslangic = QDateEdit()
        self.tarih_baslangic.setCalendarPopup(True)
        self.tarih_baslangic.setDate(QDate.currentDate().addMonths(-1))
        self.tarih_baslangic.setEnabled(False)
        tarih_row.addWidget(self.tarih_baslangic)
        
        tarih_row.addWidget(QLabel("Bitiş:"))
        self.tarih_bitis = QDateEdit()
        self.tarih_bitis.setCalendarPopup(True)
        self.tarih_bitis.setDate(QDate.currentDate())
        self.tarih_bitis.setEnabled(False)
        tarih_row.addWidget(self.tarih_bitis)
        
        self.cb_tarih_filter.stateChanged.connect(
            lambda: self.toggle_numeric_filter(self.cb_tarih_filter, [self.tarih_baslangic, self.tarih_bitis])
        )
        
        tarih_layout.addLayout(tarih_row)
        
        layout.addWidget(tarih_group)
        
        layout.addStretch()
        
        return widget
    
    def create_advanced_text_tab(self):
        """Gelişmiş metin arama tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Regex arama
        regex_group = QGroupBox("🔤 Regex (Düzenli İfade) Arama")
        regex_layout = QVBoxLayout(regex_group)
        
        self.cb_use_regex = QCheckBox("Regex kullan")
        regex_layout.addWidget(self.cb_use_regex)
        
        regex_layout.addWidget(QLabel("Regex Pattern:"))
        self.regex_input = QLineEdit()
        self.regex_input.setPlaceholderText("Örn: ^PONT.*2024$")
        self.regex_input.setEnabled(False)
        regex_layout.addWidget(self.regex_input)
        
        # Regex örnekleri
        examples = QTextEdit()
        examples.setReadOnly(True)
        examples.setMaximumHeight(120)
        examples.setText(
            "📖 Regex Örnekleri:\n"
            "^PONT.*     : PONT ile başlayan tüm JCL'ler\n"
            ".*2024$     : 2024 ile biten tüm JCL'ler\n"
            "[A-Z]{4}\\d{2}: 4 harf + 2 rakam (örn: PONT24)\n"
            "(TEST|PROD) : TEST veya PROD içeren JCL'ler"
        )
        regex_layout.addWidget(examples)
        
        self.cb_use_regex.stateChanged.connect(
            lambda: self.regex_input.setEnabled(self.cb_use_regex.isChecked())
        )
        
        layout.addWidget(regex_group)
        
        # Çoklu alan arama
        multi_group = QGroupBox("📝 Çoklu Alan Arama")
        multi_layout = QVBoxLayout(multi_group)
        
        multi_layout.addWidget(QLabel("Arama metni birden fazla alanda aransın:"))
        
        self.cb_search_jcl = QCheckBox("✓ JCL Adı")
        self.cb_search_jcl.setChecked(True)
        
        self.cb_search_sheet = QCheckBox("✓ Sheet Adı")
        self.cb_search_ekip = QCheckBox("✓ Sorumlu Ekip")
        self.cb_search_durum = QCheckBox("✓ Durum")
        
        multi_layout.addWidget(self.cb_search_jcl)
        multi_layout.addWidget(self.cb_search_sheet)
        multi_layout.addWidget(self.cb_search_ekip)
        multi_layout.addWidget(self.cb_search_durum)
        
        layout.addWidget(multi_group)
        
        layout.addStretch()
        
        return widget
    
    def create_bulk_jcl_tab(self):
        """Toplu JCL sorgulama tab'ı - Excel paste özelliği"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Açıklama
        info_label = QLabel(
            "📋 <b>Excel'den JCL Listesi Yapıştırma:</b><br>"
            "Excel'de bir kolondaki JCL adlarını kopyalayın ve aşağıdaki alana yapıştırın.<br>"
            "Her satırda bir JCL adı olmalıdır. <b>Wildcard desteği:</b> pont*, *sam, ?vsam gibi.<br>"
            "Sorgula butonuna bastığınızda <b>sağ taraftaki Arama Sonuçları</b> panelinde gösterilecek."
        )
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: #e3f2fd; padding: 10px; border-radius: 5px;")
        layout.addWidget(info_label)
        
        # JCL Girişi
        input_group = QGroupBox("JCL Listesi (Her satırda bir JCL)")
        input_layout = QVBoxLayout(input_group)
        
        self.bulk_jcl_text = QTextEdit()
        self.bulk_jcl_text.setPlaceholderText(
            "Excel'den kopyaladığınız JCL listesini buraya yapıştırın...\n\n"
            "Örnek:\n"
            "popgg001\n"
            "mag*\n"
            "*vsam\n"
            "pont*\n"
            "..."
        )
        input_layout.addWidget(self.bulk_jcl_text)
        layout.addWidget(input_group)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        
        btn_query_bulk = QPushButton("🔍 Sorgula")
        btn_query_bulk.clicked.connect(self.query_bulk_jcl)
        btn_query_bulk.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; padding: 10px;")
        
        btn_clear_bulk = QPushButton("🗑️ Temizle")
        btn_clear_bulk.clicked.connect(lambda: self.bulk_jcl_text.clear())
        
        btn_layout.addWidget(btn_query_bulk)
        btn_layout.addWidget(btn_clear_bulk)
        btn_layout.addStretch()
        
        layout.addLayout(btn_layout)
        layout.addStretch()
        
        return widget
    
    def query_bulk_jcl(self):
        """Toplu JCL sorgulama yap"""
        jcl_text = self.bulk_jcl_text.toPlainText().strip()
        
        if not jcl_text:
            QMessageBox.warning(self, "Uyarı", "Lütfen JCL listesi girin!")
            return
        
        # Satırlara ayır ve temizle
        jcl_list = [line.strip() for line in jcl_text.split('\n') if line.strip()]
        
        if not jcl_list:
            QMessageBox.warning(self, "Uyarı", "Geçerli JCL adı bulunamadı!")
            return
        
        self.logger.info(f"Toplu JCL sorgulama başlatıldı: {len(jcl_list)} JCL")
        self.status_label.setText(f"⏳ {len(jcl_list)} JCL sorgulanıyor...")
        
        # Tüm verileri al
        hatali_data = self.db_manager.get_all_hatali_isler()
        uzun_data = self.db_manager.get_all_uzun_isler()
        
        # JCL'leri sorgula (wildcard desteği ile)
        results = []
        found_count = 0
        not_found_count = 0
        
        import re
        
        for jcl in jcl_list:
            jcl_upper = jcl.upper()
            
            # Wildcard pattern oluştur
            has_wildcard = '*' in jcl_upper or '?' in jcl_upper
            if has_wildcard:
                # Wildcard'ı regex'e çevir
                pattern = jcl_upper.replace('*', '.*').replace('?', '.')
                pattern = f'^{pattern}$'  # Tam eşleşme için
            
            # Wildcard varsa: eşleşen her JCL için ayrı satır
            # Wildcard yoksa: tek satır özet
            if has_wildcard:
                matched_jcls = {}  # JCL adı -> bilgileri
                
                # Hatalı işlerde ara
                for record in hatali_data:
                    record_jcl = record.get('jcl_adi', '')
                    if re.match(pattern, record_jcl.upper()):
                        if record_jcl not in matched_jcls:
                            matched_jcls[record_jcl] = {
                                'ekip': set(),
                                'rapor': set(),
                                'ay': set()
                            }
                        if record.get('sorumlu_ekip'):
                            matched_jcls[record_jcl]['ekip'].add(record['sorumlu_ekip'])
                        matched_jcls[record_jcl]['rapor'].add('HATALI')
                        if record.get('ay'):
                            matched_jcls[record_jcl]['ay'].add(record['ay'])
                
                # Uzun işlerde ara
                for record in uzun_data:
                    record_jcl = record.get('jcl_adi', '')
                    if re.match(pattern, record_jcl.upper()):
                        if record_jcl not in matched_jcls:
                            matched_jcls[record_jcl] = {
                                'ekip': set(),
                                'rapor': set(),
                                'ay': set()
                            }
                        if record.get('sorumlu_ekip'):
                            matched_jcls[record_jcl]['ekip'].add(record['sorumlu_ekip'])
                        matched_jcls[record_jcl]['rapor'].add('UZUN')
                        if record.get('ay'):
                            matched_jcls[record_jcl]['ay'].add(record['ay'])
                
                # Her eşleşen JCL için ayrı satır ekle
                if matched_jcls:
                    for jcl_name, info in sorted(matched_jcls.items()):
                        found_count += 1
                        results.append({
                            'jcl': jcl_name,
                            'durum': '✅ Bulundu',
                            'ekip': ', '.join(sorted(info['ekip'])) if info['ekip'] else '⚠️ Ekip bilgisi yok',
                            'rapor': ', '.join(sorted(info['rapor'])),
                            'ay': ', '.join(sorted(info['ay'], reverse=True)[:3])
                        })
                else:
                    # Wildcard ile hiç eşleşme yok
                    not_found_count += 1
                    results.append({
                        'jcl': f"{jcl} (wildcard)",
                        'durum': '❌ Eşleşme yok',
                        'ekip': '-',
                        'rapor': '-',
                        'ay': '-'
                    })
            else:
                # Wildcard yok - normal tek satır göster
                found = False
                ekip_bilgisi = set()
                rapor_tipi = set()
                ay_bilgisi = set()
                
                # Hatalı işlerde ara
                for record in hatali_data:
                    if record.get('jcl_adi', '').upper() == jcl_upper:
                        found = True
                        if record.get('sorumlu_ekip'):
                            ekip_bilgisi.add(record['sorumlu_ekip'])
                        rapor_tipi.add('HATALI')
                        if record.get('ay'):
                            ay_bilgisi.add(record['ay'])
                
                # Uzun işlerde ara
                for record in uzun_data:
                    if record.get('jcl_adi', '').upper() == jcl_upper:
                        found = True
                        if record.get('sorumlu_ekip'):
                            ekip_bilgisi.add(record['sorumlu_ekip'])
                        rapor_tipi.add('UZUN')
                        if record.get('ay'):
                            ay_bilgisi.add(record['ay'])
                
                if found:
                    found_count += 1
                    results.append({
                        'jcl': jcl,
                        'durum': '✅ Bulundu',
                        'ekip': ', '.join(sorted(ekip_bilgisi)) if ekip_bilgisi else '⚠️ Ekip bilgisi yok',
                        'rapor': ', '.join(sorted(rapor_tipi)),
                        'ay': ', '.join(sorted(ay_bilgisi, reverse=True)[:3])
                    })
                else:
                    not_found_count += 1
                    results.append({
                        'jcl': jcl,
                        'durum': '❌ Bulunamadı',
                        'ekip': '-',
                        'rapor': '-',
                        'ay': '-'
                    })
        
        # Sonuçları ortak tabloya dönüştür ve göster
        formatted_results = []
        for result in results:
            formatted_results.append({
                'tur': 'TOPLU',
                'jcl_adi': result['jcl'],
                'ay': result['ay'],
                'deger': result['durum'],
                'sheet': result['rapor'],
                'ekip': result['ekip'],
                'durum': result['durum'],
                'tarih': ''
            })
        
        # Ortak sonuç tablosunu kullan
        self.populate_results(formatted_results)
        
        # Özet
        self.results_count_label.setText(
            f"📊 Toplu Sorgulama: {len(jcl_list)} JCL | "
            f"✅ Bulundu: {found_count} | "
            f"❌ Bulunamadı: {not_found_count}"
        )
        
        self.logger.info(f"Toplu sorgulama tamamlandı: {found_count} bulundu, {not_found_count} bulunamadı")
        self.status_label.setText(f"✅ Sorgulama tamamlandı: {found_count}/{len(jcl_list)} bulundu")
    
    def create_history_favorites_tab(self):
        """Geçmiş ve favoriler tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Favoriler
        fav_group = QGroupBox("⭐ Favori Aramalar")
        fav_layout = QVBoxLayout(fav_group)
        
        self.favorites_list = QListWidget()
        self.favorites_list.itemDoubleClicked.connect(self.load_favorite)
        fav_layout.addWidget(self.favorites_list)
        
        fav_buttons = QHBoxLayout()
        
        btn_load_fav = QPushButton("📂 Yükle")
        btn_load_fav.clicked.connect(self.load_favorite)
        
        btn_delete_fav = QPushButton("🗑️ Sil")
        btn_delete_fav.clicked.connect(self.delete_favorite)
        
        fav_buttons.addWidget(btn_load_fav)
        fav_buttons.addWidget(btn_delete_fav)
        
        fav_layout.addLayout(fav_buttons)
        
        layout.addWidget(fav_group)
        
        # Geçmiş
        history_group = QGroupBox("📋 Arama Geçmişi")
        history_layout = QVBoxLayout(history_group)
        
        self.history_list = QListWidget()
        self.history_list.itemDoubleClicked.connect(self.load_from_history)
        history_layout.addWidget(self.history_list)
        
        history_buttons = QHBoxLayout()
        
        btn_load_history = QPushButton("📂 Yükle")
        btn_load_history.clicked.connect(self.load_from_history)
        
        btn_clear_history = QPushButton("🗑️ Temizle")
        btn_clear_history.clicked.connect(self.clear_history)
        
        history_buttons.addWidget(btn_load_history)
        history_buttons.addWidget(btn_clear_history)
        
        history_layout.addLayout(history_buttons)
        
        layout.addWidget(history_group)
        
        # Listeleri doldur
        self.refresh_favorites_list()
        self.refresh_history_list()
        
        return widget
    
    def create_results_panel(self):
        """Sonuçlar paneli"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Başlık
        title = QLabel("📊 Arama Sonuçları")
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2196F3;")
        layout.addWidget(title)
        
        # Sonuç sayısı label
        self.results_count_label = QLabel("Sonuç bekleniyor...")
        layout.addWidget(self.results_count_label)
        
        # Sonuç tablosu
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(8)
        self.results_table.setHorizontalHeaderLabels([
            "Tür", "JCL Adı", "Ay", "Hatalı/Süre", "Sheet", "Ekip", "Durum", "Tarih"
        ])
        
        # Tablo ayarları
        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # JCL Adı
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # Sheet
        
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        layout.addWidget(self.results_table)
        
        # Export butonları
        export_buttons = QHBoxLayout()
        
        btn_export_excel = QPushButton("📊 Excel'e Aktar")
        btn_export_excel.clicked.connect(self.export_results_excel)
        
        btn_export_csv = QPushButton("📄 CSV'ye Aktar")
        btn_export_csv.clicked.connect(self.export_results_csv)
        
        export_buttons.addWidget(btn_export_excel)
        export_buttons.addWidget(btn_export_csv)
        export_buttons.addStretch()
        
        layout.addLayout(export_buttons)
        
        return widget
    
    def create_bottom_buttons(self):
        """Alt butonlar"""
        layout = QHBoxLayout()
        
        self.status_label = QLabel("Hazır")
        self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
        
        btn_close = QPushButton("❌ Kapat")
        btn_close.clicked.connect(self.close)
        
        layout.addWidget(self.status_label)
        layout.addStretch()
        layout.addWidget(btn_close)
        
        return layout
    
    def toggle_numeric_filter(self, checkbox, widgets):
        """Sayısal filtre aktif/pasif"""
        enabled = checkbox.isChecked()
        for widget in widgets:
            widget.setEnabled(enabled)
    
    def load_combo_data(self):
        """Combo box'ları doldur"""
        # Ekip ve ay verilerini yükle
        ekipler = set()
        aylar = set()
        
        hatali_data = self.db_manager.get_all_hatali_isler()
        uzun_data = self.db_manager.get_all_uzun_isler()
        
        for record in hatali_data + uzun_data:
            if record.get('sorumlu_ekip'):
                ekipler.add(record['sorumlu_ekip'])
            if record.get('ay'):
                aylar.add(record['ay'])
        
        # Ekip
        self.ekip_combo.clear()
        self.ekip_combo.addItem("Tümü")
        for ekip in sorted(ekipler):
            self.ekip_combo.addItem(ekip)
        
        # Ay
        self.ay_combo.clear()
        self.ay_combo.addItem("Tümü")
        for ay in sorted(aylar, reverse=True):
            self.ay_combo.addItem(ay)
    
    def perform_search(self):
        """Aramayı gerçekleştir"""
        self.logger.info("Gelişmiş arama başlatıldı")
        self.status_label.setText("⏳ Aranıyor...")
        self.results_table.setRowCount(0)
        
        # Arama kriterlerini topla
        criteria = self.get_search_criteria()
        
        # Arama detaylarını logla
        search_desc = criteria['jcl'] if criteria['jcl'] else "Tümü"
        if criteria['use_regex']:
            search_desc = f"Regex: {criteria['regex_pattern']}"
        self.logger.info(f"Arama kriterleri: {search_desc}, Ekip: {criteria['ekip']}, Ay: {criteria['ay']}")
        
        # Veritabanından ara
        results = self.search_database(criteria)
        
        # Sonuçları tabloya ekle
        self.populate_results(results)
        
        # Aramayı geçmişe ekle
        self.add_to_history(criteria)
        
        # Durum güncelle
        count = len(results)
        self.logger.info(f"Gelişmiş arama tamamlandı: {count} sonuç bulundu")
        self.status_label.setText(f"✅ {count} sonuç bulundu")
        self.results_count_label.setText(f"📊 Toplam {count} sonuç bulundu")
    
    def get_search_criteria(self):
        """Arama kriterlerini al"""
        criteria = {
            'jcl': self.jcl_input.text().strip(),
            'ekip': self.ekip_combo.currentText(),
            'ay': self.ay_combo.currentText(),
            'hatali_rapor': self.cb_hatali.isChecked(),
            'uzun_rapor': self.cb_uzun.isChecked(),
            'wildcard': self.cb_jcl_wildcard.isChecked(),
            'case_sensitive': self.cb_jcl_case_sensitive.isChecked(),
            'use_regex': self.cb_use_regex.isChecked(),
            'regex_pattern': self.regex_input.text().strip(),
            'search_fields': {
                'jcl': self.cb_search_jcl.isChecked(),
                'sheet': self.cb_search_sheet.isChecked(),
                'ekip': self.cb_search_ekip.isChecked(),
                'durum': self.cb_search_durum.isChecked()
            },
            'numeric_filters': {
                'hatali': {
                    'enabled': self.cb_hatali_filter.isChecked(),
                    'min': self.hatali_min.value(),
                    'max': self.hatali_max.value()
                },
                'sure': {
                    'enabled': self.cb_sure_filter.isChecked(),
                    'min': self.sure_min.value(),
                    'max': self.sure_max.value()
                },
                'tarih': {
                    'enabled': self.cb_tarih_filter.isChecked(),
                    'baslangic': self.tarih_baslangic.date().toString('yyyy-MM-dd'),
                    'bitis': self.tarih_bitis.date().toString('yyyy-MM-dd')
                }
            }
        }
        
        return criteria
    
    def search_database(self, criteria):
        """Veritabanında ara"""
        results = []
        
        # Hatalı işler
        if criteria['hatali_rapor']:
            hatali_results = self.search_hatali_isler(criteria)
            results.extend(hatali_results)
        
        # Uzun işler
        if criteria['uzun_rapor']:
            uzun_results = self.search_uzun_isler(criteria)
            results.extend(uzun_results)
        
        return results
    
    def search_hatali_isler(self, criteria):
        """Hatalı işlerde ara"""
        all_records = self.db_manager.get_all_hatali_isler()
        results = []
        
        for record in all_records:
            if self.matches_criteria(record, criteria, 'HATALI'):
                results.append({
                    'tur': 'HATALI',
                    'jcl_adi': record.get('jcl_adi', ''),
                    'ay': record.get('ay', ''),
                    'deger': record.get('hatali_sayi_ay', 0),
                    'sheet': record.get('sheet_adi', ''),
                    'ekip': record.get('sorumlu_ekip', ''),
                    'durum': record.get('durum', ''),
                    'tarih': record.get('yuklenme_tarihi', '')
                })
        
        return results
    
    def search_uzun_isler(self, criteria):
        """Uzun işlerde ara"""
        all_records = self.db_manager.get_all_uzun_isler()
        results = []
        
        for record in all_records:
            if self.matches_criteria(record, criteria, 'UZUN'):
                results.append({
                    'tur': 'UZUN',
                    'jcl_adi': record.get('jcl_adi', ''),
                    'ay': record.get('ay', ''),
                    'deger': record.get('sure_dk', 0),
                    'sheet': record.get('sheet_adi', ''),
                    'ekip': record.get('sorumlu_ekip', ''),
                    'durum': record.get('durum', ''),
                    'tarih': record.get('yuklenme_tarihi', '')
                })
        
        return results
    
    def matches_criteria(self, record, criteria, rapor_tipi):
        """Kayıt kriterlere uyuyor mu?"""
        import re
        
        # JCL arama - Regex ÖNCE kontrol edilmeli
        if criteria['use_regex'] and criteria['regex_pattern']:
            # REGEX ARAMA
            jcl_adi = record.get('jcl_adi', '')
            try:
                pattern = criteria['regex_pattern']
                if not criteria['case_sensitive']:
                    if not re.search(pattern, jcl_adi, re.IGNORECASE):
                        return False
                else:
                    if not re.search(pattern, jcl_adi):
                        return False
            except re.error as e:
                self.logger.warning(f"Geçersiz regex pattern: {pattern}, Hata: {e}")
                return False
        elif criteria['jcl']:
            # NORMAL/WILDCARD ARAMA
            jcl_text = criteria['jcl']
            jcl_adi = record.get('jcl_adi', '')
            
            if True:  # Normal arama bloğu
                # Normal arama
                search_terms = [term.strip() for term in jcl_text.split(',')]
                match_found = False
                
                for term in search_terms:
                    if criteria['wildcard']:
                        # Wildcard desteği
                        pattern = term.replace('*', '.*').replace('?', '.')
                        if not criteria['case_sensitive']:
                            if re.search(pattern, jcl_adi, re.IGNORECASE):
                                match_found = True
                                break
                        else:
                            if re.search(pattern, jcl_adi):
                                match_found = True
                                break
                    else:
                        # Exact match
                        if not criteria['case_sensitive']:
                            if term.lower() in jcl_adi.lower():
                                match_found = True
                                break
                        else:
                            if term in jcl_adi:
                                match_found = True
                                break
                
                if not match_found:
                    return False
        
        # Ekip filtresi
        if criteria['ekip'] != 'Tümü':
            if record.get('sorumlu_ekip') != criteria['ekip']:
                return False
        
        # Ay filtresi
        if criteria['ay'] != 'Tümü':
            if record.get('ay') != criteria['ay']:
                return False
        
        # Sayısal filtreler
        numeric = criteria['numeric_filters']
        
        # Hatalı sayı aralığı
        if rapor_tipi == 'HATALI' and numeric['hatali']['enabled']:
            hatali_sayi = record.get('hatali_sayi_ay', 0)
            if not (numeric['hatali']['min'] <= hatali_sayi <= numeric['hatali']['max']):
                return False
        
        # Süre aralığı
        if rapor_tipi == 'UZUN' and numeric['sure']['enabled']:
            sure = record.get('sure_dk', 0)
            if not (numeric['sure']['min'] <= sure <= numeric['sure']['max']):
                return False
        
        return True
    
    def populate_results(self, results):
        """Sonuçları tabloya ekle"""
        self.results_table.setRowCount(len(results))
        
        for row, result in enumerate(results):
            # Tür
            tur_item = QTableWidgetItem(result['tur'])
            if result['tur'] == 'HATALI':
                tur_item.setBackground(QColor("#ffebee"))
            else:
                tur_item.setBackground(QColor("#e3f2fd"))
            self.results_table.setItem(row, 0, tur_item)
            
            # Diğer kolonlar
            self.results_table.setItem(row, 1, QTableWidgetItem(str(result['jcl_adi'])))
            self.results_table.setItem(row, 2, QTableWidgetItem(str(result['ay'])))
            self.results_table.setItem(row, 3, QTableWidgetItem(str(result['deger'])))
            self.results_table.setItem(row, 4, QTableWidgetItem(str(result['sheet'])))
            self.results_table.setItem(row, 5, QTableWidgetItem(str(result['ekip'])))
            self.results_table.setItem(row, 6, QTableWidgetItem(str(result['durum'])))
            self.results_table.setItem(row, 7, QTableWidgetItem(str(result['tarih'])))
    
    def reset_criteria(self):
        """Kriterleri sıfırla"""
        self.jcl_input.clear()
        self.ekip_combo.setCurrentIndex(0)
        self.ay_combo.setCurrentIndex(0)
        self.cb_hatali.setChecked(True)
        self.cb_uzun.setChecked(True)
        self.cb_jcl_wildcard.setChecked(True)
        self.cb_jcl_case_sensitive.setChecked(False)
        self.cb_use_regex.setChecked(False)
        self.regex_input.clear()
        
        # Sayısal filtreler
        self.cb_hatali_filter.setChecked(False)
        self.cb_sure_filter.setChecked(False)
        self.cb_tarih_filter.setChecked(False)
        
        self.status_label.setText("✅ Kriterler sıfırlandı")
    
    def save_current_search(self):
        """Mevcut aramayı kaydet"""
        from PyQt5.QtWidgets import QInputDialog
        
        name, ok = QInputDialog.getText(self, "Aramayı Kaydet", "Favori adı:")
        
        if ok and name:
            criteria = self.get_search_criteria()
            self.favorites[name] = {
                'criteria': criteria,
                'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            self.save_favorites()
            self.refresh_favorites_list()
            
            QMessageBox.information(self, "Başarılı", f"✅ '{name}' favorilere eklendi!")
    
    def load_favorite(self):
        """Favoriyi yükle"""
        current_item = self.favorites_list.currentItem()
        if current_item:
            name = current_item.text().split(' - ')[0]
            if name in self.favorites:
                self.load_criteria(self.favorites[name]['criteria'])
                self.status_label.setText(f"✅ '{name}' favorisi yüklendi")
    
    def delete_favorite(self):
        """Favoriyi sil"""
        current_item = self.favorites_list.currentItem()
        if current_item:
            name = current_item.text().split(' - ')[0]
            if name in self.favorites:
                reply = QMessageBox.question(
                    self, "Onay", f"'{name}' favorisini silmek istediğinize emin misiniz?",
                    QMessageBox.Yes | QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    del self.favorites[name]
                    self.save_favorites()
                    self.refresh_favorites_list()
                    self.status_label.setText(f"🗑️ '{name}' silindi")
    
    def add_to_history(self, criteria):
        """Geçmişe ekle"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        jcl_text = criteria['jcl'] if criteria['jcl'] else "Tümü"
        
        self.search_history.insert(0, {
            'criteria': criteria,
            'timestamp': timestamp,
            'description': f"{jcl_text} - {timestamp}"
        })
        
        # Maksimum 20 arama geçmişi
        self.search_history = self.search_history[:20]
        
        self.save_search_history()
        self.refresh_history_list()
    
    def load_from_history(self):
        """Geçmişten yükle"""
        current_row = self.history_list.currentRow()
        if current_row >= 0 and current_row < len(self.search_history):
            history_item = self.search_history[current_row]
            self.load_criteria(history_item['criteria'])
            self.status_label.setText("✅ Geçmişten yüklendi")
    
    def clear_history(self):
        """Geçmişi temizle"""
        reply = QMessageBox.question(
            self, "Onay", "Tüm arama geçmişini silmek istediğinize emin misiniz?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.search_history = []
            self.save_search_history()
            self.refresh_history_list()
            self.status_label.setText("🗑️ Geçmiş temizlendi")
    
    def load_criteria(self, criteria):
        """Kriterleri form'a yükle"""
        self.jcl_input.setText(criteria.get('jcl', ''))
        
        ekip_index = self.ekip_combo.findText(criteria.get('ekip', 'Tümü'))
        if ekip_index >= 0:
            self.ekip_combo.setCurrentIndex(ekip_index)
        
        ay_index = self.ay_combo.findText(criteria.get('ay', 'Tümü'))
        if ay_index >= 0:
            self.ay_combo.setCurrentIndex(ay_index)
        
        self.cb_hatali.setChecked(criteria.get('hatali_rapor', True))
        self.cb_uzun.setChecked(criteria.get('uzun_rapor', True))
        self.cb_jcl_wildcard.setChecked(criteria.get('wildcard', True))
        self.cb_jcl_case_sensitive.setChecked(criteria.get('case_sensitive', False))
        
        # Regex
        self.cb_use_regex.setChecked(criteria.get('use_regex', False))
        self.regex_input.setText(criteria.get('regex_pattern', ''))
        
        # Sayısal filtreler
        numeric = criteria.get('numeric_filters', {})
        
        if 'hatali' in numeric:
            self.cb_hatali_filter.setChecked(numeric['hatali']['enabled'])
            self.hatali_min.setValue(numeric['hatali']['min'])
            self.hatali_max.setValue(numeric['hatali']['max'])
        
        if 'sure' in numeric:
            self.cb_sure_filter.setChecked(numeric['sure']['enabled'])
            self.sure_min.setValue(numeric['sure']['min'])
            self.sure_max.setValue(numeric['sure']['max'])
    
    def refresh_favorites_list(self):
        """Favoriler listesini yenile"""
        self.favorites_list.clear()
        for name, data in self.favorites.items():
            self.favorites_list.addItem(f"{name} - {data['date']}")
    
    def refresh_history_list(self):
        """Geçmiş listesini yenile"""
        self.history_list.clear()
        for item in self.search_history:
            self.history_list.addItem(item['description'])
    
    def load_search_history(self):
        """Arama geçmişini yükle"""
        try:
            if os.path.exists(self.search_history_file):
                with open(self.search_history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return []
    
    def save_search_history(self):
        """Arama geçmişini kaydet"""
        try:
            os.makedirs(os.path.dirname(self.search_history_file), exist_ok=True)
            with open(self.search_history_file, 'w', encoding='utf-8') as f:
                json.dump(self.search_history, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Geçmiş kaydetme hatası: {e}")
    
    def load_favorites(self):
        """Favorileri yükle"""
        try:
            if os.path.exists(self.favorites_file):
                with open(self.favorites_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return {}
    
    def save_favorites(self):
        """Favorileri kaydet"""
        try:
            os.makedirs(os.path.dirname(self.favorites_file), exist_ok=True)
            with open(self.favorites_file, 'w', encoding='utf-8') as f:
                json.dump(self.favorites, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Favoriler kaydetme hatası: {e}")
    
    def export_results_excel(self):
        """Sonuçları Excel'e aktar"""
        if self.results_table.rowCount() == 0:
            QMessageBox.warning(self, "Uyarı", "Dışa aktarılacak sonuç yok!")
            return
        
        self.logger.info("Gelişmiş arama: Excel export başlatıldı")
        from PyQt5.QtWidgets import QFileDialog
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel'e Aktar", "", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                row_count = self.results_table.rowCount()
                self.logger.info(f"Excel export: {row_count} kayıt aktarılıyor")
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Arama Sonuçları"
                
                # Başlıklar
                headers = []
                for col in range(self.results_table.columnCount()):
                    headers.append(self.results_table.horizontalHeaderItem(col).text())
                
                ws.append(headers)
                
                # Veriler
                for row in range(self.results_table.rowCount()):
                    row_data = []
                    for col in range(self.results_table.columnCount()):
                        item = self.results_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    ws.append(row_data)
                
                # Stil
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                wb.save(file_path)
                
                self.logger.info(f"Excel export başarılı: {file_path}")
                QMessageBox.information(self, "Başarılı", f"✅ Sonuçlar Excel'e aktarıldı:\n{file_path}")
                
            except Exception as e:
                self.logger.error(f"Excel export hatası: {e}")
                QMessageBox.critical(self, "Hata", f"Excel aktarma hatası:\n{e}")
    
    def export_results_csv(self):
        """Sonuçları CSV'ye aktar"""
        if self.results_table.rowCount() == 0:
            QMessageBox.warning(self, "Uyarı", "Dışa aktarılacak sonuç yok!")
            return
        
        self.logger.info("Gelişmiş arama: CSV export başlatıldı")
        from PyQt5.QtWidgets import QFileDialog
        import csv
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "CSV'ye Aktar", "", "CSV Files (*.csv)"
        )
        
        if file_path:
            try:
                row_count = self.results_table.rowCount()
                self.logger.info(f"CSV export: {row_count} kayıt aktarılıyor")
                
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    
                    # Başlıklar
                    headers = []
                    for col in range(self.results_table.columnCount()):
                        headers.append(self.results_table.horizontalHeaderItem(col).text())
                    writer.writerow(headers)
                    
                    # Veriler
                    for row in range(self.results_table.rowCount()):
                        row_data = []
                        for col in range(self.results_table.columnCount()):
                            item = self.results_table.item(row, col)
                            row_data.append(item.text() if item else "")
                        writer.writerow(row_data)
                
                self.logger.info(f"CSV export başarılı: {file_path}")
                QMessageBox.information(self, "Başarılı", f"✅ Sonuçlar CSV'ye aktarıldı:\n{file_path}")
                
            except Exception as e:
                self.logger.error(f"CSV export hatası: {e}")
                QMessageBox.critical(self, "Hata", f"CSV aktarma hatası:\n{e}")
